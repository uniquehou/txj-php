from django.shortcuts import render
from AdminManage.models import Organization
from FreeCourse.models import FreeCourse
from django.http import HttpResponse, HttpResponseRedirect, StreamingHttpResponse
from django.urls import reverse
import json
import os

def login(request):
    if request.method == "GET":
        return render(request, 'AdminManage/login.html')
    else:
        try:
            organic = Organization.objects.get(name=request.POST['name'])
            if organic.password == request.POST['password']:
                request.session['id'] = organic.id
                request.session['name'] = organic.name
                data = {"status": "1", "image": organic.image, "name": organic.name}    # 登录成功
                if organic.organic_name:
                    data['name'] = organic.organic_name
                return HttpResponse(json.dumps(data)) 
            else:
                return HttpResponse(json.dumps({"status": "2"}))    # 密码不正确
        except DoesNotExist:
            return HttpResponse(json.dumps({"status": "0"}))        # 此用户不存在
        except:
            return HttpResponse(json.dumps({"status": "3"}))

def regis(request):
    if request.method == "GET":
        return render(request, 'AdminManage/regis.html')
    else:
        name = request.POST["name"]
        organic_name = request.POST["organic_name"]
        password = request.POST["password"]
        organic = Organization.objects.create(name=name, password=password, organic_name=organic_name)
        FreeCourse.objects.create(organic_id=organic)
        return HttpResponse("1")        # 注册成功

def index(request):
    organic = Organization.objects.get(id=request.session.get("id"))
    return render(request, 'AdminManage/index.html')
    
def info_manage(request):
    if request.method == "GET":
        return render(request, "AdminManage/info_manage.html")
    else:
        organic = Organization.objects.get(id=request.session.get("id"))
        organic.organic_name = request.POST['organic_name']
        organic.save()
        return HttpResponse('1')

def password_modify(request):
    if request.method == "GET":
        return render(request, 'AdminManage/password_modify.html')
    else:
        organic = Organization.objects.get(id=request.session.get("id"))
        if organic.password == request.POST['old_password']:
            organic.password = request.POST['new_password']
            organic.save()
            return HttpResponse("1")       # 修改成功
        else:
            return HttpResponse("2")       # 密码不正确

def password_find(request):
    pass

def free_course_manage(request):
    if not request.GET.get('open'):
        organic = Organization.objects.get(id=request.session.get('id'))
        data = {"free_course_open": organic.free_course_open, "id": organic.id}
        return render(request, 'AdminManage/free_course_manage.html', data)
    else:
        organic = Organization.objects.get(id=request.session.get('id'))
        organic.free_course_open = request.GET.get('open')
        organic.save()
        return HttpResponse(organic.free_course_open)
            

def free_course_show(request):
    organic = FreeCourse.objects.get(organic_id=request.session.get("id"))
    data = {'data': organic.data, 'id':request.session.get('id')}
    return render(request, 'AdminManage/free_course_show.html', data)

def file_iterator(file_name, chunk_size=512):
    with open(file_name) as f:
        while True:
            c = f.read(chunk_size)
            if c:
                yield c
            else:
                break 

def download(request):
    # the_file_name = "txj.xls"
    # response = StreamingHttpResponse(file_iterator(the_file_name))
    # response['Content-Type'] = 'application/octet-stream'
    # response['Content-Disposition'] = 'attachment;filename="{0}"'.format(the_file_name)
    # return response

    from openpyxl import Workbook

    free_course = json.loads(FreeCourse.objects.get(organic_id=request.session.get('id')).data)
    file_name = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))) + '/media/FreeCourse/', request.session.get('name') + '.xls')
    
    wb = Workbook()
    ws = wb.active
    # table head
    weeks = ['周一', '周二', '周三', '周四', '周五', '周六', '周日']
    for i in range(2, 9):
        ws.cell(row=1, column=i).value = weeks[i-2]
    ws.cell(row=2, column=1).value = "上午"
    ws.cell(row=4, column=1).value = "下午"
    ws.cell(row=6, column=1).value = "晚自习"
    # data
    for col in range(7):
        for row in range(1, 6):
            course = "course_" + str(col*5+row)
            ws.cell(row=row+1, column=col+2).value = free_course[course]
    wb.save(file_name)
    return HttpResponseRedirect('/txj/media/FreeCourse/%s' % request.session.get('name') + '.xls')


def exit(request):
    if request.session.get("id"):
        del request.session['id']
    if request.session.get("name"):
        del request.session['name']
    return HttpResponse("1")


